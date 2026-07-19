import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

output_path = r"C:\Users\zhanglin\Desktop\app-coobohaoham9\会员健康信息采集表.xlsx"

wb = openpyxl.Workbook()

# ============ 样式定义 ============
title_font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
subtitle_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
label_font = Font(name="微软雅黑", size=11, bold=True)
normal_font = Font(name="微软雅黑", size=11)
small_font = Font(name="微软雅黑", size=10)
red_font = Font(name="微软雅黑", size=10, color="C00000")

thin_side = Side(style="thin", color="BFBFBF")
medium_side = Side(style="medium", color="7F7F7F")
border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
section_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
pink_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# ============ Sheet 1: 采集表 ============
ws1 = wb.active
ws1.title = "健康信息采集表"
ws1.freeze_panes = "A6"

# 列宽
ws1.column_dimensions["A"].width = 8
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 30
ws1.column_dimensions["G"].width = 18

# 标题
ws1.merge_cells("A1:G1")
cell = ws1["A1"]
cell.value = "会员健康信息采集表"
cell.font = title_font
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = header_fill
ws1.row_dimensions[1].height = 36

# 副标题
ws1.merge_cells("A2:G2")
cell = ws1["A2"]
cell.value = "请认真填写，信息越完整，调理方案越精准；舌象请参照《拍照参考》页拍摄。"
cell.font = small_font
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell.fill = light_fill
ws1.row_dimensions[2].height = 24

# 基础信息区
ws1.merge_cells("A3:G3")
cell = ws1["A3"]
cell.value = "一、基础信息"
cell.font = subtitle_font
cell.alignment = Alignment(horizontal="left", vertical="center")
cell.fill = section_fill
ws1.row_dimensions[3].height = 24

basic_info = [
    ["性别", "", "年龄", "", "生肖属相", ""],
    ["身高（cm）", "", "体重（kg）", "", "联系电话", ""],
]
row = 4
for r in basic_info:
    for i, v in enumerate(r):
        col = i + 1
        c = ws1.cell(row=row, column=col, value=v)
        c.font = label_font if i % 2 == 0 else normal_font
        c.alignment = Alignment(horizontal="center" if i % 2 == 0 else "left", vertical="center")
        c.border = border
        if i % 2 == 0:
            c.fill = light_fill
    row += 1

# 舌照区
ws1.merge_cells(f"A{row}:G{row}")
cell = ws1.cell(row=row, column=1, value="二、舌象照片（请按《拍照参考》页要求拍摄，并粘贴照片于此）")
cell.font = subtitle_font
cell.alignment = Alignment(horizontal="left", vertical="center")
cell.fill = section_fill
ws1.row_dimensions[row].height = 24
row += 1

ws1.merge_cells(f"A{row}:G{row+3}")
cell = ws1.cell(row=row, column=1, value="")
cell.border = border
cell.fill = pink_fill
for rr in range(row, row+4):
    ws1.row_dimensions[rr].height = 48
row += 4

# 问诊区
ws1.merge_cells(f"A{row}:G{row}")
cell = ws1.cell(row=row, column=1, value="三、健康问诊（请逐题详细填写）")
cell.font = subtitle_font
cell.alignment = Alignment(horizontal="left", vertical="center")
cell.fill = section_fill
ws1.row_dimensions[row].height = 24
row += 1

# 表头
headers = ["序号", "项目", "具体问题", "填写区", "", "", ""]
for i, h in enumerate(headers):
    c = ws1.cell(row=row, column=i+1, value=h)
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = header_fill
    c.border = header_border
ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
ws1.row_dimensions[row].height = 28
row += 1

questions = [
    ("1", "饮食偏好", "平时偏爱哪些食物？如水果、冷饮、甜食、辛辣刺激、油炸、油腻、海鲜、奶茶等。"),
    ("2", "食欲消化", "胃口如何？有无反酸、打嗝、嗳气、腹胀、不消化、口苦、口臭？"),
    ("3", "二便情况", "小便颜色/频率？大便一天几次？干燥/便秘？黏腻难冲？腹泻？排便是否费力？"),
    ("4", "女性/男性专项", "女性：月经是否规律？痛经（小腹/后腰）？经前胸胀/头痛？经血颜色（黑/暗红/褐）？量？天数？生育/流产？\n男性：是否吸烟饮酒？夜尿几次？夫妻生活是否正常？有无腰酸/乏力？"),
    ("5", "全身状态", "是否容易疲乏、犯困、浑身酸痛？有无气短/胸闷/心慌？易感冒？怕冷/怕热？上热下寒？忽冷忽热？"),
    ("6", "情绪状态", "是否容易焦虑、抑郁、易怒、悲伤、恐惧、健忘、失眠多梦、精神紧张？"),
    ("7", "睡眠情况", "入睡快吗？一般几点睡？多梦吗？易惊醒？睡不沉/早醒？白天困倦？"),
    ("8", "疾病/用药/手术", "已确诊疾病及确诊时间？重大外伤或手术史？近期是否服用中西药？请说明。"),
    ("9", "疫苗情况", "是否接种新冠疫苗？接种时间？接种后身体有无明显变化？"),
]

for idx, item, q in questions:
    start_row = row
    c1 = ws1.cell(row=row, column=1, value=idx)
    c1.font = label_font
    c1.alignment = Alignment(horizontal="center", vertical="top")
    c1.border = border
    c1.fill = light_fill

    c2 = ws1.cell(row=row, column=2, value=item)
    c2.font = label_font
    c2.alignment = Alignment(horizontal="center", vertical="top")
    c2.border = border
    c2.fill = light_fill

    c3 = ws1.cell(row=row, column=3, value=q)
    c3.font = small_font
    c3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c3.border = border

    ws1.merge_cells(start_row=row, start_column=4, end_row=row+2, end_column=7)
    c4 = ws1.cell(row=row, column=4, value="")
    c4.border = border
    c4.fill = pink_fill
    c4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for rr in range(row, row+3):
        ws1.row_dimensions[rr].height = 36
    row += 3

# 提交区
row += 1
ws1.merge_cells(f"A{row}:G{row}")
cell = ws1.cell(row=row, column=1, value="")
cell.border = border
ws1.row_dimensions[row].height = 12
row += 1

ws1.merge_cells(f"A{row}:D{row}")
cell = ws1.cell(row=row, column=1, value="会员签名：")
cell.font = label_font
cell.alignment = Alignment(horizontal="left", vertical="center")
cell.border = border
ws1.merge_cells(f"E{row}:G{row}")
cell = ws1.cell(row=row, column=5, value="填写日期：      年    月    日")
cell.font = label_font
cell.alignment = Alignment(horizontal="left", vertical="center")
cell.border = border
ws1.row_dimensions[row].height = 28

# 全部列统一边框
for col in range(1, 8):
    for r in range(1, row+1):
        ws1.cell(row=r, column=col).border = border

# ============ Sheet 2: 拍照参考 ============
ws2 = wb.create_sheet("拍照参考")
ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 70

ws2.merge_cells("A1:B1")
cell = ws2["A1"]
cell.value = "舌象拍摄参考"
cell.font = title_font
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = header_fill
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:B2")
cell = ws2["A2"]
cell.value = "舌头自然伸展，不要过于用力，放松点；不要僵硬，舌尖微微过下嘴唇。"
cell.font = small_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
cell.fill = pink_fill
ws2.row_dimensions[2].height = 36

shooting_items = [
    ("1", "选择光线柔和、自然的拍照环境，避免强光或背光。"),
    ("2", "舌体自然伸出，不要用力、不要卷曲，避免舌头僵硬。"),
    ("3", "相机/手机不要开美颜、滤镜，需提交原图。"),
    ("4", "从舌的正前方拍摄，最好能看清舌根；不能用俯视或侧视角度。"),
    ("5", "请别人帮你拍摄，能清楚看到舌根部。自拍容易角度不准。"),
    ("6", "照片要露出鼻尖、下巴，鼻唇沟与下巴正中成直线。"),
    ("7", "建议拍摄两张：正面舌象（苔色、舌质）+ 舌下静脉（卷起舌尖）。"),
    ("", "舌面脏腑对应：舌尖→心肺，舌中→脾胃，舌根→肾，舌边→肝胆。"),
]

row = 3
for num, txt in shooting_items:
    c1 = ws2.cell(row=row, column=1, value=num)
    c1.font = label_font
    c1.alignment = Alignment(horizontal="center", vertical="top")
    c1.border = border
    c1.fill = light_fill

    c2 = ws2.cell(row=row, column=2, value=txt)
    c2.font = normal_font
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c2.border = border

    ws2.row_dimensions[row].height = 32
    row += 1

ws2.merge_cells(f"A{row}:B{row+3}")
cell = ws2.cell(row=row, column=1, value="【舌象照片粘贴区】\n可贴：正面舌象 / 舌下静脉照片")
cell.font = normal_font
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell.border = border
ws2.row_dimensions[row].height = 36
ws2.row_dimensions[row+1].height = 36
ws2.row_dimensions[row+2].height = 36
ws2.row_dimensions[row+3].height = 36

wb.save(output_path)
print(f"已生成：{output_path}")
