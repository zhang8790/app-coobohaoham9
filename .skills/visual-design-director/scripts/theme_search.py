#!/usr/bin/env python3
"""
Theme template lookup — first-priority path when user picks a named template.

Flow:
1. Read template_name (from user message <AESTHETIC_TEMPLATE>...</AESTHETIC_TEMPLATE>)
2. If template_name in {"自动", "auto", "automatic"} or empty → exit 1 (fallback to Step 2)
3. Open data/theme.xlsx, filter rows where:
   - data.is_published_to_prod == True
   - data.title == template_name
   - data.support_app_types contains app_type (case-insensitive substring)
4. If --output is given, write data.context to that path; always print to stdout
5. Exit codes:
   - 0: hit, context printed to stdout (and written to --output if given)
   - 1: miss (auto/empty, template not found, or app_type incompatible — reason on stderr)
"""

import argparse
import sys
from pathlib import Path

import openpyxl

THEME_XLSX = Path(__file__).resolve().parent.parent / "data" / "theme.xlsx"

AUTO_VALUES = {"", "自动", "auto", "automatic"}

# PRD app_type → xlsx support_app_types lookup value.
# xlsx only has: Web, Mobile App, MiniProgram.
# H5/Tool/Questionnaire/Others all use Web templates.
_APP_TYPE_TO_XLSX = {
    "web": "web",
    "h5": "web",
    "tool": "web",
    "questionnaire": "web",
    "others": "web",
    "mobile app": "mobileapp",
    "mini program": "miniprogram",
}


def _normalize(s: str) -> str:
    """Lowercase + strip + remove all whitespace.

    Handles xlsx variants like "MiniProgram" vs PRD's "Mini Program";
    also makes case-insensitive.
    """
    return "".join(str(s or "").split()).lower()


def _matches_app_type(support_field: str, app_type: str) -> bool:
    """Check whether app_type matches support_app_types.

    Maps PRD app_type (Web/H5/Tool/Questionnaire/Others/Mobile App/Mini Program)
    to xlsx lookup value (web/mobileapp/miniprogram) via _APP_TYPE_TO_XLSX,
    then checks substring match against normalized support_field.
    """
    if not support_field or not app_type:
        return False
    haystack = _normalize(support_field)
    needle = _APP_TYPE_TO_XLSX.get(app_type.strip().lower())
    if not needle:
        return False
    return needle in haystack


def _dir_state(output_path) -> str:
    return "ready" if output_path else "skipped"


def _theme_status(written: bool, output_path) -> str:
    next_step = "no_read" if written else "content_below"
    return (
        f"written={str(written).lower()} | dir={_dir_state(output_path)} | "
        f"{next_step}"
    )


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument(
        "--template-name",
        required=True,
        help='User-selected template name from <AESTHETIC_TEMPLATE>. '
             'Pass "自动"/"auto" or empty to skip.',
    )
    p.add_argument(
        "--app-type",
        required=True,
        help="PRD app_type, e.g. Web / H5 / Mobile App / Mini Program / Tool / Questionnaire.",
    )
    p.add_argument(
        "--output",
        default=None,
        help="If given, write DESIGN.md content to this path on hit (parent dirs created automatically).",
    )
    args = p.parse_args()

    template_name = (args.template_name or "").strip()
    app_type = (args.app_type or "").strip()

    if args.output:
        out_parent = Path(args.output).parent
        out_parent.mkdir(parents=True, exist_ok=True)
        print(f"[theme_search] 输出目录已就绪: {out_parent}", file=sys.stderr)

    if template_name in AUTO_VALUES:
        print(f"[theme_search] template_name={template_name!r} is auto/empty — skip", file=sys.stderr)
        return 1

    if not THEME_XLSX.exists():
        print(f"[theme_search] theme.xlsx not found at {THEME_XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(THEME_XLSX, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        idx_prod = headers.index("data.is_published_to_prod")
        idx_title = headers.index("data.title")
        idx_support = headers.index("data.support_app_types")
        idx_context = headers.index("data.context")
    except ValueError as e:
        print(f"[theme_search] required column missing: {e}", file=sys.stderr)
        return 1

    title_seen = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx_prod]:
            continue
        if (row[idx_title] or "").strip() != template_name:
            continue
        title_seen = True
        if not _matches_app_type(row[idx_support], app_type):
            continue
        context = row[idx_context]
        if context:
            written = False
            if args.output:
                out = Path(args.output)
                out.write_text(context, encoding="utf-8")
                print(f"[theme_search] written to {out}", file=sys.stderr)
                written = True
            print("## 已命中模板")
            print(f"**{template_name}** | {_theme_status(written, args.output)}")
            print("")
            print("--- DESIGN.md 当前内容开始 ---")
            print(context)
            print("--- DESIGN.md 当前内容结束 ---")
            return 0

    if title_seen:
        print(
            f"[theme_search] template {template_name!r} exists but no row supports app_type={app_type!r}",
            file=sys.stderr,
        )
    else:
        print(
            f"[theme_search] template {template_name!r} not found in published themes",
            file=sys.stderr,
        )
    return 1


if __name__ == "__main__":
    sys.exit(main())
