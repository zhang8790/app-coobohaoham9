#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Design System Generator — 2-path auto-search + full 3+3+3 BM25 candidates.

Paths (internal, triggered automatically):
  T  theme.xlsx BM25 score >= THEME_MIN  → write DESIGN.md from theme.context,
                                           output validation header + anti-slop
  C  no theme hit                        → product / style(3) / color(3) / font(3) + scenario
                                           + anti-slop
"""

import re
import sys
import colorsys
from pathlib import Path
from core import BM25, search

ANTI_SLOP_FILE = (
    Path(__file__).parent.parent / "data" / "anti-slop.md"
)

THEME_XLSX = Path(__file__).parent.parent / "data" / "theme.xlsx"
THEME_MIN = 4.0
THEME_CONFIDENT_MIN = 5.0

# theme.xlsx 是上游 CMS 导出（不可本地改）。description/context 含通用词
# （如"数字色彩""高端大气"）会误召回；标题是强视觉信号，但只在 query
# 与标题有非泛词重合时加权，避免"功能主义"命中"神秘主义"。
THEME_TITLE_WEIGHT = 5.0
THEME_DESC_WEIGHT = 1.0
THEME_CONTEXT_WEIGHT = 0.25
THEME_EXACT_BONUS = 8.0

_THEME_GENERIC_TOKENS = frozenset({
    "主义", "风格", "界面", "工具", "专业", "现代", "简洁", "高端",
    "数字", "系统", "平台", "应用", "页面", "主题", "视觉", "设计",
    "功能", "管理", "数据", "问卷", "网页", "网站", "色调", "传统",
    "温暖", "清新",
    "style", "design", "modern", "clean", "tool", "app", "system",
    "platform", "page", "theme", "visual", "professional", "data",
    "digital",
})
_THEME_STRONG_TOKENS = frozenset({
    "像素", "喜庆", "国风", "暗黑", "极简", "森系", "手写", "蓝图",
    "工程", "终端", "奢华", "治愈", "霓虹", "赛博", "复古", "水墨",
    "笔记", "记本", "双色", "几何", "网格", "点亮", "立体", "红色",
    "白色", "午夜", "梦幻", "自然", "文艺", "神秘",
    "pixel", "festive", "dark", "notebook", "blueprint", "terminal",
    "opulent", "macaron", "natural", "mystical", "brutal", "duotone",
    "geometric", "neumorphic", "editorial", "retro", "cyberpunk",
    "minimal",
})
_THEME_FALLBACK_TOKENS = _THEME_STRONG_TOKENS - frozenset({
    "professional", "minimal",
})

# 候选字体 URL 前缀；搜索结果只给文件名，agent 写 DESIGN.md 时拼接此前缀。
# 改前缀只需改这一处。
FONT_URL_PREFIX = "https://resource-static.bj.bcebos.com/fonts-skill/"

_APP_TYPE_NEEDLE = {
    "web": "web", "h5": "web", "tool": "web",
    "questionnaire": "web", "others": "web",
    "mobile app": "mobileapp", "mini program": "miniprogram",
}

PLATFORM_MOBILE = "mobile"
PLATFORM_WEB = "web"
PLATFORM_ANY = "any"
VALID_PLATFORMS = frozenset({PLATFORM_MOBILE, PLATFORM_WEB, PLATFORM_ANY})

# requirementType → 检索 platform（直接映射，agent 无需传 --platform）。
# Mobile App / Mini Program → mobile（移动端样式 + APP 字体过滤）；H5 → any（不过滤）；
# Web / Tool / Questionnaire / Others 及未知 → web。
_APP_TYPE_PLATFORM = {
    "mobile app": PLATFORM_MOBILE,
    "mini program": PLATFORM_MOBILE,
    "h5": PLATFORM_ANY,
    "web": PLATFORM_WEB,
    "tool": PLATFORM_WEB,
    "questionnaire": PLATFORM_WEB,
    "others": PLATFORM_WEB,
}


def _platform_from_app_type(app_type: str) -> str:
    return _APP_TYPE_PLATFORM.get((app_type or "").strip().lower(), PLATFORM_WEB)


# ─── theme auto-search ────────────────────────────────────────────────────────

def _theme_tokens(text: str) -> set:
    return set(BM25().tokenize(text))


def _meaningful_title_overlap(query: str, title: str) -> set:
    overlap = _theme_tokens(query) & _theme_tokens(title)
    return {
        token for token in overlap
        if token in _THEME_STRONG_TOKENS
        or token not in _THEME_GENERIC_TOKENS
    }


def _bm25_scores(query: str, documents: list) -> dict:
    bm25 = BM25()
    bm25.fit(documents)
    return dict(bm25.score(query))

def _auto_theme_search(query: str, app_type: str) -> dict:
    """BM25 search on published theme.xlsx rows filtered by app_type."""
    if not THEME_XLSX.exists():
        return {"hit": False}
    needle = _APP_TYPE_NEEDLE.get((app_type or "").strip().lower(), "web")
    try:
        import openpyxl  # lazy: 仅休眠回退用；沙盒可能无此库，缺失即降级未命中
        wb = openpyxl.load_workbook(THEME_XLSX, data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        i_prod = headers.index("data.is_published_to_prod")
        i_title = headers.index("data.title")
        i_desc = headers.index("data.description")
        i_support = headers.index("data.support_app_types")
        i_ctx = headers.index("data.context")
    except Exception:
        return {"hit": False}

    title_docs, desc_docs, context_docs, meta = [], [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[i_prod]:
            continue
        support = "".join(str(row[i_support] or "").split()).lower()
        if needle and needle not in support:
            continue
        title = str(row[i_title] or "")
        ctx = str(row[i_ctx] or "")
        if not title or not ctx:
            continue
        title_docs.append(title)
        desc_docs.append(str(row[i_desc] or ""))
        context_docs.append(ctx)
        meta.append({"title": title, "context": ctx})

    if not meta:
        return {"hit": False}

    title_scores = _bm25_scores(query, title_docs)
    desc_scores = _bm25_scores(query, desc_docs)
    context_scores = _bm25_scores(query, context_docs)
    ranked = []
    for idx, item in enumerate(meta):
        overlap = _meaningful_title_overlap(query, item["title"])
        title_score = title_scores.get(idx, 0) if overlap else 0
        score = (
            THEME_TITLE_WEIGHT * title_score
            + THEME_DESC_WEIGHT * desc_scores.get(idx, 0)
            + THEME_CONTEXT_WEIGHT * context_scores.get(idx, 0)
        )
        if query.strip() == item["title"].strip():
            score += THEME_EXACT_BONUS
        ranked.append((idx, score, overlap))
    ranked.sort(key=lambda x: x[1], reverse=True)

    if not ranked or ranked[0][1] < THEME_MIN:
        return {"hit": False}
    idx, score, overlap = ranked[0]
    if score < THEME_CONFIDENT_MIN and not overlap:
        return {"hit": False}

    return {"hit": True, "score": score, **meta[idx]}


def _strong_theme_query(query: str) -> str:
    """Extract only strong visual tokens for a conservative fallback search."""
    tokens = [
        token for token in BM25().tokenize(query)
        if token in _THEME_FALLBACK_TOKENS
    ]
    return " ".join(dict.fromkeys(tokens))


# ─── anti-slop ────────────────────────────────────────────────────────────────

def _load_anti_slop() -> list:
    if not ANTI_SLOP_FILE.exists():
        return []
    lines = ANTI_SLOP_FILE.read_text(encoding="utf-8").splitlines()
    rules = []
    for line in lines:
        line = line.strip()
        if line.startswith(
            ("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.")
        ):
            parts = line.split("**", 2)
            if len(parts) >= 3:
                rules.append(parts[1].strip())
            else:
                rules.append(line[3:].strip())
    return rules


# ─── theme Path T output ──────────────────────────────────────────────────────

def _dir_state(output_path: str) -> str:
    return "ready" if output_path else "skipped"


def _theme_status(written: bool, output_path: str) -> str:
    next_step = "no_read" if written else "content_below"
    return (
        f"written={str(written).lower()} | dir={_dir_state(output_path)} | "
        f"{next_step}"
    )


def _format_theme_hit(theme: dict, anti_slop: list, written: bool, output_path: str) -> str:
    lines = [
        "## 已命中模板",
        f"**{theme['title']}** [score={theme['score']:.2f}] | {_theme_status(written, output_path)}",
    ]
    # 两种分支都把 context 放进模型上下文：写盘时避免模型回头 read 文件，未写盘时供其写入。
    lines += [
        "",
        "--- DESIGN.md 当前内容开始 ---",
        theme["context"],
        "--- DESIGN.md 当前内容结束 ---",
        "",
    ]
    if anti_slop:
        lines.append("## Anti-Slop Rules")
        lines.extend(f"- {r}" for r in anti_slop)
        lines.append("")
    return "\n".join(lines)


# ─── scenario helper ──────────────────────────────────────────────────────────

def _get_scenario(query: str, platform: str):
    sc_hits = search(query, "scenario", 1).get("results") or []
    scenario = sc_hits[0] if sc_hits else None
    _WEB_ONLY = {
        "Portfolio / Showcase", "Official Website",
        "Blog / Content", "Admin Console",
    }
    if platform == PLATFORM_MOBILE and (
        not scenario or scenario.get("Scenario") in _WEB_ONLY
    ):
        fb = search("mobile app native touch", "scenario", 1)
        scenario = (fb.get("results") or [None])[0] or scenario
    return scenario


# ─── platform filtering ───────────────────────────────────────────────────────

def _filter_by_platform(styles: list, platform: str) -> list:
    if platform == PLATFORM_WEB:
        return [
            s for s in styles
            if s.get("Type", "").strip() != "Mobile"
            and "(Mobile)" not in s.get("Style Category", "")
        ]
    if platform == PLATFORM_MOBILE:
        return [
            s for s in styles
            if s.get("Type", "").strip() == "Mobile"
            or "(Mobile)" in s.get("Style Category", "")
        ]
    return styles


# ─── color diversity ──────────────────────────────────────────────────────────

COOL_DARK = frozenset({"neutral", "blue", "cyan", "green"})


def _color_bucket(hex_color: str) -> str:
    hex_color = hex_color.lstrip("#")
    if len(hex_color) != 6:
        return "unknown"
    r = int(hex_color[0:2], 16) / 255
    g = int(hex_color[2:4], 16) / 255
    b = int(hex_color[4:6], 16) / 255
    h, s, v = colorsys.rgb_to_hsv(r, g, b)
    if s < 0.15 or v < 0.2:
        return "neutral"
    hue = h * 360
    if hue < 30 or hue >= 330:
        return "red"
    elif hue < 75:
        return "orange"
    elif hue < 150:
        return "green"
    elif hue < 210:
        return "cyan"
    elif hue < 270:
        return "blue"
    return "purple"


def _ensure_color_diversity(colors: list, query: str) -> list:
    """Inject a warm palette when top-3 lack hue variety."""
    if len(colors) < 3:
        return colors
    buckets = [_color_bucket(c.get("Primary", "#000")) for c in colors]
    from collections import Counter
    bucket_counts = Counter(buckets)
    most_common_count = bucket_counts.most_common(1)[0][1]
    all_cool = all(b in COOL_DARK for b in buckets)
    if most_common_count < 2 and not all_cool:
        return colors
    extended = search(query, "color", 10)
    extended_colors = extended.get("results") or []
    for candidate in extended_colors[3:]:
        cb = _color_bucket(candidate.get("Primary", "#000"))
        if cb not in COOL_DARK and cb not in buckets:
            colors[-1] = candidate
            return colors
    warm_fallback = search("warm vibrant creative", "color", 5)
    warm_colors = warm_fallback.get("results") or []
    for candidate in warm_colors:
        cb = _color_bucket(candidate.get("Primary", "#000"))
        if cb not in COOL_DARK:
            colors[-1] = candidate
            return colors
    return colors


# ─── font platform / language filters ───────────────────────────────────────

def _font_platform_ok(font: dict, platform: str) -> bool:
    """Mobile (native app / mini program) only accepts 平台=APP fonts;
    web / any (incl. H5) is unrestricted."""
    if platform != PLATFORM_MOBILE:
        return True
    return str(font.get("平台", "")).strip().upper() == "APP"



def _font_covers_user_lang(font: dict, user_lang: str) -> bool:
    """Stricter than _font_renders_cjk: matches font coverage to user language.
    Chinese apps require 简/繁 fonts; Japanese apps accept 日文 or 简/繁 (which
    usually cover the common Japanese kanji set). Note: 'japanese' in a search
    query describes aesthetic style, not the user language — user_lang comes
    from LANGUAGE_SETTINGS only."""
    font_lang = str(font.get("language", ""))
    low = (user_lang or "").lower()
    if "japanese" in low or low == "ja":
        return ("日" in font_lang) or ("简" in font_lang) or ("繁" in font_lang) or ("中" in font_lang)
    if "korean" in low or low in ("ko", "ko-kr"):
        return "韩" in font_lang
    # Chinese (simplified / traditional) — exclude Japanese-only fonts
    return ("简" in font_lang) or ("繁" in font_lang) or ("中" in font_lang)


def _lang_needs_cjk(lang: str) -> bool:
    """App language (lang_user) requires a CJK-capable body font.
    lang_user values from agenthub are English words: 'Chinese', 'Japanese', 'English'.
    Also handles BCP-47 tags and CJK characters as fallback."""
    low = (lang or "").lower()
    return (
        "zh" in low or "chinese" in low
        or "japanese" in low or low == "ja"
        or "korean" in low or low in ("ko", "ko-kr")
        or bool(re.search(r"[一-鿿ぁ-んァ-ン가-힣]", lang or ""))
    )


# ─── font coverage ──────────────────────────────────────────────────────────

def _font_search_depth(platform: str) -> int:
    """补齐池深度。CJK body 同样是少数派：Latin-heavy 的 query 下它常排在 20 名外，
    web/any 旧的浅检索（20）会让 _ensure_font_coverage 注入兜底够不着，导致中文应用
    候选无可用正文（实测 id-79）。语料仅 ~126 行，全量 BM25 成本可忽略，故各平台都全量。
    注意：_select_fonts 的 top-3 取 `[:3]`，与 depth 无关，故正常选字行为不变，
    仅扩大注入兜底的可达范围。"""
    return 200


def _ensure_font_coverage(
    fonts: list, query: str,
    platform: str = PLATFORM_WEB, cjk_app: bool = False,
    lang: str = "",
) -> list:
    """Top-3 must hold >=1 heading-capable and >=1 body-capable font.
    Body must satisfy the app language: a CJK app cannot use a Latin-only
    font for body text. Chinese apps reject Japanese-only fonts. Replacements
    are drawn only from platform-eligible fonts."""
    if len(fonts) < 2:
        return fonts

    def body_ok(f):
        if f.get("Usage", "") not in ("body", "both"):
            return False
        if not cjk_app:
            return True
        return _font_covers_user_lang(f, lang)

    def heading_ok(f):
        return f.get("Usage", "") in ("heading", "both")

    if any(heading_ok(f) for f in fonts) and any(body_ok(f) for f in fonts):
        return fonts

    extended = [
        c for c in (search(query, "font", _font_search_depth(platform)).get("results") or [])
        if _font_platform_ok(c, platform)
    ]

    def _inject(role_ok, keep_ok):
        """补一个 role_ok 字体，且不挤掉唯一提供 keep_ok 的槽位。
        优先替换不满足 keep_ok 的槽，避免 heading/body 两个稀缺角色互相踩踏；
        无安全槽时退回末位。"""
        if any(role_ok(f) for f in fonts):
            return
        cand = next(
            (c for c in extended if role_ok(c) and c not in fonts), None
        )
        if cand is None:
            return
        keepers = [i for i, f in enumerate(fonts) if keep_ok(f)]
        if len(keepers) == 1:
            idx = next(
                (i for i in range(len(fonts)) if i not in keepers),
                len(fonts) - 1,
            )
        else:
            idx = len(fonts) - 1
        fonts[idx] = cand

    # body 对 CJK 应用更稀缺，先补 body；补 heading 时不覆盖唯一的 body 槽。
    _inject(body_ok, keep_ok=heading_ok)
    _inject(heading_ok, keep_ok=body_ok)
    return fonts


# ─── motion dial ─────────────────────────────────────────────────────────────

def _motion_label(mb) -> str:
    val = int(mb)
    if val <= 2:
        return "ultra-minimal (loading indicators only)"
    if val <= 4:
        return "functional (subtle state transitions)"
    if val <= 6:
        return "moderate (entrance + hover)"
    if val <= 8:
        return "expressive (choreographed scroll reveals)"
    return "immersive (full choreography, motion as narrative)"


# ─── text helpers ─────────────────────────────────────────────────────────────

def _clean_tech(text: str) -> str:
    text = re.sub(
        r"(box-shadow|text-shadow|border-radius|backdrop-filter)"
        r":\s*[^;,]+[;,]?",
        "",
        text,
    )
    text = re.sub(r"rgba?\([^)]+\)", "", text)
    text = re.sub(r"clamp\([^)]*\)", "", text)
    text = re.sub(
        r"(font-size|font-weight|letter-spacing|line-height):\s*[^;,]*[;,]?",
        "",
        text,
    )
    text = re.sub(r"#[0-9A-Fa-f]{3,8}\b", "", text)
    text = re.sub(r"\d+px\b", "", text)
    text = re.sub(r"\d+pt\b", "", text)
    text = re.sub(r"\d+(\.\d+)?(rem|vw|vh|em)\b", "", text)
    text = re.sub(r"\s*[;,](\s*[;,])+", ";", text)
    text = re.sub(r"\s{2,}", " ", text).strip()
    return text.strip("; ,").strip()


def _truncate(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text
    cut = text[:limit]
    for sep in ("; ", ". ", ", "):
        idx = cut.rfind(sep)
        if idx > limit // 2:
            return cut[:idx]
    return cut.rstrip() + "…"


# def _strip_hex(text: str) -> str:
#     text = re.sub(r"#[0-9A-Fa-f]{3,8}\b", "", text)
#     return re.sub(r"\s{2,}", " ", text).strip()

# ─── candidate formatter ─────────────────────────────────────────────────────


def _font_fit_text(font: dict) -> str:
    fit = str(font.get("Best_For", "")).strip()
    if fit:
        return _truncate(fit, 90)
    return _truncate(str(font.get("Mood_Keywords", "")).strip(), 90)


def _format_fonts(fonts: list, lang: str = "") -> list:
    """Render compact Font Candidates (fit + weights + 文件名 + language)."""
    out = [
        "## Font Candidates",
        f"font_prefix={FONT_URL_PREFIX}",
    ]
    _low = (lang or "").lower()
    use_cn = "zh" in _low or "chinese" in _low
    for f in fonts:
        en_name = f.get("Font_Name_EN") or f.get("Font_Family", "?")
        cn_name = f.get("Font_Name_CN", "")
        name = (cn_name if cn_name and cn_name != en_name else en_name) if use_cn else en_name
        cat = f.get("Category", "")
        family = f.get("Font_Family", en_name)
        template = f.get("CDN_URL_Template", "")
        url = f.get("CDN_URL", "")
        fit = _font_fit_text(f)
        weights = f.get("Weights", "")
        usage = f.get("Usage", "")
        line = f"- {name} ({cat})"
        if family != name:
            line += f" fam={family}"
        if usage:
            line += f" [{usage}]"
        if fit:
            line += f" | fit={fit}"
        if weights:
            line += f" | w={weights}"
        # 只给文件名；agent 拼前缀。Path().name 兼容文件名或整段 URL。
        fname = template if (template and "{weight}" in template) else url
        if fname:
            line += f" | file={Path(fname).name}"
        language = f.get("language", "")
        if language:
            line += f" | lang={language}"
        out.append(line)
    out.append("")
    return out


def _mark_latin_heading_only(fonts: list, lang: str) -> list:
    """CJK 应用里，渲染不了用户语言的字体（西文 only / 异语 CJK）不能做正文。
    把其展示 Usage 由 body/both 压成 heading，避免 agent 把 `[both]` 误当成
    可用正文（实测漏洞：HarmonyOS Sans 西文标 both 被选作中文 Body）。
    只改展示、不删字体——纯西文字体仍可用于拉丁标题。"""
    for f in fonts:
        if not _font_covers_user_lang(f, lang) and f.get("Usage") in ("body", "both"):
            f["Usage"] = "heading"
    return fonts


def _select_fonts(font_query: str, platform: str, lang: str) -> list:
    """选字体：平台过滤 → 取 3 → 覆盖补齐（heading/body + 中日文 Body）。
    主流程 generate 与 fonts_only 共用此逻辑，单一来源。"""
    if platform not in VALID_PLATFORMS:
        platform = PLATFORM_WEB
    # 中文/日文应用判定：lang_user 指明，或检索词含 CJK 字符。
    cjk_app = _lang_needs_cjk(lang) or bool(re.search(r"[一-鿿]", font_query))
    font_pool = [
        f for f in (search(font_query, "font", _font_search_depth(platform)).get("results") or [])
        if _font_platform_ok(f, platform)
    ]
    fonts = font_pool[:3]
    fonts = _ensure_font_coverage(fonts, font_query, platform, cjk_app, lang)
    if cjk_app:
        fonts = _mark_latin_heading_only(fonts, lang)
    return fonts


def fonts_only(query: str, platform: str = PLATFORM_MOBILE, lang: str = "") -> str:
    """仅输出 Font Candidates 段（Mobile App 分支自定配色/布局，只取 CDN 字体池）。"""
    plat = platform if platform in VALID_PLATFORMS else PLATFORM_MOBILE
    return "\n".join(_format_fonts(_select_fonts(query, plat, lang), lang))


def _format_candidates(
    scenario,
    styles: list,
    colors: list,
    fonts: list,
    anti_slop: list,
    output_path: str = "",
    lang: str = "",
) -> str:
    lines = [
        f"## Result: theme=miss | dir={_dir_state(output_path)} | no_retry",
        "",
    ]

    # Scenario Brief
    if scenario:
        sc_name = scenario.get("Scenario", "")
        sc_rules = scenario.get("Layout_Rules", "")
        mb = scenario.get("Motion_Baseline", "")
        ac = scenario.get("Animation_Constraint", "")
        parts = []
        if sc_rules:
            parts.append(f"layout={sc_rules}")
        if mb:
            parts.append(f"motion={mb}/10 {_motion_label(mb)}")
        if ac:
            parts.append(f"anim={ac}")
        lines.append("## Scenario")
        lines.append(f"{sc_name} | {' | '.join(parts)}" if parts else sc_name)
        lines.append("")

    # Style Reference
    lines.append("## Style Reference")
    for s in styles:
        name = s.get("Style Category", "?")
        elements = _clean_tech(s.get("Signature_Elements", ""))
        effects = _clean_tech(s.get("Effects & Animation", ""))
        fx_parts = []
        if effects:
            el_lower = elements.lower()
            fx_items = re.split(r"[;,]\s*(?![^()]*\))", effects)
            for item in fx_items:
                item = item.strip()
                if not item:
                    continue
                core = re.sub(r"\([^)]*\)", "", item).strip().lower()
                if len(core) >= 4 and core in el_lower:
                    continue
                fx_parts.append(item)
        dna = elements
        if fx_parts:
            extra = "; ".join(fx_parts[:3])
            dna = f"{dna}; {extra}" if dna else extra
        dna = _truncate(dna, 160)
        lines.append(f"- {name}: {dna}")
    lines.append("")

    # Color Candidates
    lines.append("## Color Candidates")
    lines.append("legend: P/on primary; A/on accent; S secondary; BG/FG/M/B bg/fg/muted/border")
    color_fields = [
        ("Primary", "P"), ("On Primary", "on"),
        ("Accent", "A"), ("On Accent", "on"),
        ("Secondary", "S"),
        ("Background", "BG"), ("Foreground", "FG"),
        ("Muted", "M"), ("Border", "B"),
    ]
    for color in colors:
        cpt = color.get("Product Type", "")
        parts = []
        for key, label in color_fields:
            val = color.get(key, "")
            if val:
                parts.append(f"{label} {val}")
        if parts:
            label = "-"
            if cpt:
                label += f" {cpt}:"
            lines.append(f"{label} {' | '.join(parts)}")
    lines.append("")

    # Font Candidates
    lines.extend(_format_fonts(fonts, lang))

    # Anti-Slop Rules
    if anti_slop:
        lines.append("## Anti-Slop Rules")
        for rule in anti_slop:
            lines.append(f"- {rule}")
        lines.append("")

    return "\n".join(lines)


# ─── main entry point ─────────────────────────────────────────────────────────

def generate(
    query: str,
    platform: str = "",
    lang: str = "",
    app_type: str = "",
    output_path: str = "",
    theme_query: str = "",
) -> str:
    """Search and return design system candidates / auto-write theme content.

    本地化（去 MCP）：先按 app_type 过滤做 theme.xlsx BM25 自动检索，命中即把
    theme.context 写盘并返回校验头（Path T）。未命中分两支：
      - Mobile App：自配色/布局，仅出字体候选（见 mobile_app.md）。
      - 其余平台：完整 3+3+3 候选（Path C）。

    Args:
        query:        English keywords (3-8 words) for style/color/font BM25.
        platform:     'web' | 'mobile' | 'any'；省略时由 app_type 自动映射
                      （_platform_from_app_type）。仅作可选覆盖。
        lang:         User language from LANGUAGE_SETTINGS. Drives CJK-capable
                      body-font enforcement for Chinese/Japanese apps.
        app_type:     requirementType from PRDAgent result（驱动 theme 过滤 +
                      platform 映射 + 未命中分支）。
        output_path:  命中 theme 时写 DESIGN.md 内容到此路径（Path T）。
        theme_query:  Workflow-required native-language visual theme phrase.
                      Empty falls back to `query` only for compatibility.
    """
    if platform not in VALID_PLATFORMS:
        platform = _platform_from_app_type(app_type)

    if output_path:
        out_parent = Path(output_path).parent
        out_parent.mkdir(parents=True, exist_ok=True)
        print(f"[design] 输出目录已就绪: {out_parent}", file=sys.stderr)

    anti_slop = _load_anti_slop()

    # ── Path T：本地 theme 自动检索（替代 MCP），命中自动写盘 ──────────────────
    theme = _auto_theme_search(theme_query.strip() or query, app_type)
    if not theme["hit"] and theme_query.strip():
        fallback_query = _strong_theme_query(query)
        if fallback_query:
            theme = _auto_theme_search(fallback_query, app_type)
    if theme["hit"]:
        written = False
        if output_path:
            Path(output_path).write_text(
                theme["context"], encoding="utf-8"
            )
            written = True
        return _format_theme_hit(theme, anti_slop, written, output_path)

    # ── Mobile App 未命中：自配色/布局，只取字体候选 ──────────────────────────
    if app_type.strip().lower() == "mobile app":
        return "\n".join(_format_fonts(_select_fonts(query, platform, lang), lang))

    # ── Path C: no theme, full 3+3+3 BM25 ───────────────────────────────────
    prod = (search(query, "product", 1).get("results") or [{}])[0]
    scenario = _get_scenario(query, platform)

    if platform == PLATFORM_ANY:
        styles = search(query, "style", 3).get("results") or []
    elif platform == PLATFORM_MOBILE:
        # Mobile styles are a minority of the corpus; a plain top-N search
        # gets crowded out by web styles, leaving <3 after filtering. The
        # token "mobile" matches Type=Mobile so every mobile style scores,
        # and the content words still rank them — guarantees 3 candidates.
        styles = _filter_by_platform(
            search(f"{query} mobile", "style", 30).get("results") or [],
            platform,
        )[:3]
    else:
        styles = _filter_by_platform(
            search(query, "style", 10).get("results") or [], platform
        )[:3]

    prod_type = prod.get("Product Type", "")
    color_query = f"{prod_type} {query}" if prod_type else query
    colors = search(color_query, "color", 3).get("results") or []
    colors = _ensure_color_diversity(colors, query)

    font_query = (
        f"{query} "
        + " ".join(s.get("Style Category", "") for s in styles[:3])
    ).strip()
    fonts = _select_fonts(font_query, platform, lang)

    return _format_candidates(
        scenario=scenario,
        styles=styles,
        colors=colors,
        fonts=fonts,
        anti_slop=anti_slop,
        output_path=output_path,
        lang=lang,
    )
